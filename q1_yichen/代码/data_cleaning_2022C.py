#!/usr/bin/env python3
"""2022 C题玻璃文物数据清洗脚本。

运行示例：
python data_cleaning_2022C.py \
  --input 2022Cdata.xlsx \
  --output 2022Cdata_最终清洗版.xlsx

依赖：pandas、numpy、openpyxl

输出工作表：
1. 表单1/表单2/表单3：保留原始数据；
2. 表单1_清洗：统一编号、补充颜色缺失标记和风化编码；
3. 表单2_清洗：关联文物属性、判断样本有效性并归一化；
4. 模型数据：仅保留有效样本，供后续统计建模使用；
5. 质量检查：记录各项清洗结果，便于复核。
"""

from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_SHEETS = ("表单1", "表单2", "表单3")
VALID_SUM_MIN = 85.0
VALID_SUM_MAX = 105.0


def normalize_relic_id(value: object) -> str:
    """从“03部位1”等文本中提取文物编号，并统一为两位字符。"""
    if pd.isna(value):
        return ""
    match = re.match(r"\s*(\d+)", str(value))
    if match is None:
        return ""
    return f"{int(match.group(1)):02d}"


def normalize_sample_name(value: object) -> str:
    """普通数字采样点补齐前导0，带中文说明的名称保持原样。"""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return f"{int(float(text)):02d}"
    return text


def identify_sample_type(sample_name: str) -> str:
    if "未风化点" in sample_name:
        return "未风化点"
    if "严重风化点" in sample_name:
        return "严重风化点"
    if "部位" in sample_name:
        return "不同部位"
    return "普通采样点"


def clean_form1(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()
    df["文物编号"] = df["文物编号"].map(normalize_relic_id)

    for column in ("纹饰", "类型", "表面风化"):
        df[column] = df[column].astype("string").str.strip()

    color = df["颜色"].astype("string").str.strip()
    color_missing = color.isna() | color.eq("")
    df["清洗颜色"] = color.mask(color_missing, "未知")
    df["风化编码"] = df["表面风化"].map({"无风化": 0, "风化": 1})
    df["颜色缺失"] = color_missing.astype(int)

    id_ok = df["文物编号"].str.fullmatch(r"\d{2}", na=False)
    id_unique = ~df["文物编号"].duplicated(keep=False)
    pattern_ok = df["纹饰"].isin(["A", "B", "C"])
    type_ok = df["类型"].isin(["高钾", "铅钡"])
    weather_ok = df["表面风化"].isin(["风化", "无风化"])
    df["数据检查"] = np.where(
        id_ok & id_unique & pattern_ok & type_ok & weather_ok,
        "通过",
        "需检查",
    )
    return df


def clean_form2(
    raw: pd.DataFrame, form1_clean: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    df = raw.copy()
    sample_column = "文物采样点"
    chemistry_columns = [column for column in df.columns if column != sample_column]

    df[sample_column] = df[sample_column].map(normalize_sample_name)
    numeric = df[chemistry_columns].apply(pd.to_numeric, errors="coerce")
    df[chemistry_columns] = numeric

    df["文物编号"] = df[sample_column].map(normalize_relic_id)
    df["采样点类型"] = df[sample_column].map(identify_sample_type)

    relic_info = form1_clean.set_index("文物编号")
    df["玻璃类型"] = df["文物编号"].map(relic_info["类型"])
    df["文物整体风化"] = df["文物编号"].map(relic_info["表面风化"])

    # 题目中特别标明“未风化点”的采样点按无风化处理；其余沿用文物整体状态。
    df["采样点实际风化"] = np.where(
        df["采样点类型"].eq("未风化点"),
        "无风化",
        df["文物整体风化"],
    )

    zero_filled = numeric.fillna(0.0)
    total = zero_filled.sum(axis=1)
    valid = total.between(VALID_SUM_MIN, VALID_SUM_MAX, inclusive="both")
    df["原始成分总和"] = total.round(2)
    df["数据有效性"] = np.where(valid, "有效", "无效")
    df["未检出成分数"] = (numeric.isna() | numeric.eq(0)).sum(axis=1)

    normalized = zero_filled.div(total.replace(0, np.nan), axis=0).mul(100)
    normalized = normalized.where(valid, np.nan)
    normalized_columns = []
    for column in chemistry_columns:
        new_column = f"{column}_归一化"
        normalized_columns.append(new_column)
        df[new_column] = normalized[column]
    df["归一化合计"] = df[normalized_columns].sum(axis=1, min_count=1)

    if df[["玻璃类型", "文物整体风化"]].isna().any().any():
        bad = df.loc[
            df[["玻璃类型", "文物整体风化"]].isna().any(axis=1), sample_column
        ].tolist()
        raise ValueError(f"以下采样点无法与表单1匹配：{bad}")

    model_columns = [
        sample_column,
        "文物编号",
        "采样点类型",
        "玻璃类型",
        "采样点实际风化",
        "原始成分总和",
        "未检出成分数",
        *normalized_columns,
        "归一化合计",
    ]
    model = df.loc[valid, model_columns].copy()
    model = model.rename(
        columns={
            sample_column: "采样点",
            "采样点实际风化": "实际风化",
            "未检出成分数": "未检出数",
        }
    )
    return df, model, normalized_columns


def build_quality_report(
    form1: pd.DataFrame,
    form2: pd.DataFrame,
    model: pd.DataFrame,
    normalized_columns: list[str],
) -> pd.DataFrame:
    group_counts = (
        model.groupby(["玻璃类型", "实际风化"], observed=False)
        .size()
        .to_dict()
    )
    normalized_bad = int(
        (~np.isclose(model[normalized_columns].sum(axis=1), 100.0, atol=1e-6)).sum()
    )

    checks = [
        ("表单1文物数", len(form1), 58),
        ("高钾文物数", int(form1["类型"].eq("高钾").sum()), 18),
        ("铅钡文物数", int(form1["类型"].eq("铅钡").sum()), 40),
        ("风化文物数", int(form1["表面风化"].eq("风化").sum()), 34),
        ("无风化文物数", int(form1["表面风化"].eq("无风化").sum()), 24),
        ("颜色缺失数", int(form1["颜色缺失"].sum()), 4),
        ("表单1异常行数", int(form1["数据检查"].ne("通过").sum()), 0),
        ("表单2采样点数", len(form2), 69),
        ("有效采样点数", int(form2["数据有效性"].eq("有效").sum()), 67),
        ("无效采样点数", int(form2["数据有效性"].eq("无效").sum()), 2),
        ("模型数据行数", len(model), 67),
        ("高钾-无风化", int(group_counts.get(("高钾", "无风化"), 0)), 12),
        ("高钾-风化", int(group_counts.get(("高钾", "风化"), 0)), 6),
        ("铅钡-无风化", int(group_counts.get(("铅钡", "无风化"), 0)), 23),
        ("铅钡-风化", int(group_counts.get(("铅钡", "风化"), 0)), 26),
        ("归一化合计异常数", normalized_bad, 0),
    ]
    report = pd.DataFrame(checks, columns=["检查项", "实际值", "预期值"])
    report["结果"] = np.where(report["实际值"].eq(report["预期值"]), "通过", "需检查")

    invalid = form2.loc[
        form2["数据有效性"].eq("无效"), ["文物采样点", "原始成分总和"]
    ]
    invalid_text = "；".join(
        f"{row['文物采样点']}（合计{row['原始成分总和']:.2f}%）"
        for _, row in invalid.iterrows()
    )
    report.loc[len(report)] = ["无效样本", invalid_text, "保留但不进入模型", "已处理"]
    report.loc[len(report)] = [
        "清洗规则",
        "空白成分按未检出处理；85%≤总和≤105%为有效；有效样本按行归一化到100%",
        "—",
        "已执行",
    ]
    return report


def write_workbook(
    source: Path,
    output: Path,
    form1: pd.DataFrame,
    form2: pd.DataFrame,
    model: pd.DataFrame,
    report: pd.DataFrame,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == output.resolve():
        raise ValueError("输出文件不能覆盖原始文件，请指定新的输出路径。")

    # 先复制原文件，因此原始表及其格式会被保留。
    shutil.copy2(source, output)
    with pd.ExcelWriter(
        output, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        form1.to_excel(writer, sheet_name="表单1_清洗", index=False)
        form2.to_excel(writer, sheet_name="表单2_清洗", index=False)
        model.to_excel(writer, sheet_name="模型数据", index=False)
        report.to_excel(writer, sheet_name="质量检查", index=False)

    format_workbook(output)


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="0F6B78")
    header_font = Font(color="FFFFFF", bold=True)
    invalid_fill = PatternFill("solid", fgColor="F4CCCC")

    for sheet_name in ("表单1_清洗", "表单2_清洗", "模型数据", "质量检查"):
        sheet = workbook[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 30
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = min(max(max(map(len, values), default=0) + 2, 10), 24)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    form1_sheet = workbook["表单1_清洗"]
    form1_sheet.column_dimensions["A"].width = 10
    for cell in form1_sheet["A"][1:]:
        cell.number_format = "00"

    form2_sheet = workbook["表单2_清洗"]
    headers = {cell.value: cell.column for cell in form2_sheet[1]}
    if "数据有效性" in headers:
        letter = get_column_letter(headers["数据有效性"])
        form2_sheet.conditional_formatting.add(
            f"{letter}2:{letter}{form2_sheet.max_row}",
            CellIsRule(operator="equal", formula=['"无效"'], fill=invalid_fill),
        )
    for header, index in headers.items():
        if header in {"原始成分总和", "归一化合计"} or str(header).endswith("_归一化"):
            for row in range(2, form2_sheet.max_row + 1):
                form2_sheet.cell(row=row, column=index).number_format = "0.0000"

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="清洗2022年C题玻璃文物数据")
    parser.add_argument("--input", required=True, type=Path, help="原始xlsx文件")
    parser.add_argument("--output", required=True, type=Path, help="清洗后xlsx文件")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"找不到原始文件：{args.input}")

    excel = pd.ExcelFile(args.input)
    missing_sheets = [sheet for sheet in SOURCE_SHEETS if sheet not in excel.sheet_names]
    if missing_sheets:
        raise ValueError(f"原始文件缺少工作表：{missing_sheets}")

    raw1 = pd.read_excel(args.input, sheet_name="表单1")
    raw2 = pd.read_excel(args.input, sheet_name="表单2")
    form1 = clean_form1(raw1)
    form2, model, normalized_columns = clean_form2(raw2, form1)
    report = build_quality_report(form1, form2, model, normalized_columns)
    write_workbook(args.input, args.output, form1, form2, model, report)

    invalid_samples = form2.loc[
        form2["数据有效性"].eq("无效"), "文物采样点"
    ].tolist()
    print(f"清洗完成：{args.output}")
    print(f"有效样本：{len(model)}；无效样本：{invalid_samples}")


if __name__ == "__main__":
    main()
